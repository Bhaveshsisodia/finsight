"""
FPI NSDL — Node Functions Only
================================
Drop these into your existing LangGraph setup.
Each function signature:  fn(state: FPIState) -> dict
"""

from __future__ import annotations
import os, time, logging, calendar
from datetime import date
from typing import TypedDict
import requests
import pandas as pd
from bs4 import BeautifulSoup

log = logging.getLogger("fpi_nodes")

# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  STATE SCHEMA                                                                ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
class FPIState(TypedDict, total=False):
    # inputs
    today:              date
    output_dir:         str
    # detect_dates
    report_dates:       list[date]
    date_urls:          dict[str, str]
    dates_checked:      list[str]
    # fetch_reports
    raw_html:           dict[str, str]          # {date_str: html}
    fetch_errors:       list[str]
    # parse_tables
    raw_dataframes:     dict[str, list]         # {date_str: list of 98-col rows}
    parse_errors:       list[str]
    # extract_data
    sector_data:        dict[str, list[dict]]   # {date_str: [{sector fields}]}
    sector_totals:      dict[str, dict]         # {date_str: {grand totals}}
    extract_errors:     list[str]
    # generate_insights
    insights:           list[dict]
    summary:            dict
    buying_sectors:     list[dict]
    selling_sectors:    list[dict]
    top_movers:         list[dict]
    # save_reports
    saved_files:        list[str]
    excel_path:         str
    save_errors:        list[str]


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  CONSTANTS                                                                   ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
_BASE_URL = (
    "https://www.fpi.nsdl.co.in/web/StaticReports/"
    "Fortnightly_Sector_wise_FII_Investment_Data/"
    "FIIInvestSector_{ds}.html"
)
_HTTP_HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-IN,en;q=0.9",
    "Referer":         "https://www.fpi.nsdl.co.in/",
}
_MONTH = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
          7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}

# Confirmed 98-col table layout (from debug output)
_COL_SECTOR = 1
_BLOCKS     = {"AUC_F15": 2, "F1": 26, "F2": 50, "AUC_F28": 74}
_SUB        = {"Equity":0,"DebtGL":1,"DebtVRR":2,"DebtFAR":3,
               "Hybrid":4,"MF":5,"AIF":10,"Total":11}
_USD_OFFSET = 12   # USD block = INR block + 12

_SIGNAL_COLORS = {
    "STRONG BUY":  ("006400","CCFFCC"),
    "BUY":         ("006400","E8F5E9"),
    "MILD BUY":    ("856404","FFFFF0"),
    "MILD SELL":   ("7D3C00","FFF3E0"),
    "SELL":        ("CC0000","FFE5E5"),
    "STRONG SELL": ("8B0000","FFCCCC"),
}


# ── private helpers ────────────────────────────────────────────────────────────
def _ds(d: date) -> str:
    return f"{_MONTH[d.month]}{d.day:02d}{d.year}"

def _url(d: date) -> str:
    return _BASE_URL.format(ds=_ds(d))

def _num(val) -> float:
    try:    return float(str(val).replace(",","").strip())
    except: return 0.0

def _signal(v: float) -> str:
    if   v >=  1000: return "STRONG BUY"
    elif v >=   100: return "BUY"
    elif v >=     0: return "MILD BUY"
    elif v >=  -100: return "MILD SELL"
    elif v >= -1000: return "SELL"
    else:            return "STRONG SELL"

def _cidx(block: str, sub: str) -> int:
    return _BLOCKS[block] + _SUB[sub]

def _mf(row: list, block: str) -> float:
    base = _BLOCKS[block]
    return sum(_num(row[base + i]) for i in range(5, 10))


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  NODE 1 — detect_dates                                                       ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
def detect_dates(state: FPIState) -> dict:
    """
    Walk backwards from today, HEAD-check each fortnightly candidate URL,
    return the 2 most recent dates that have live reports.

    Reads  : state["today"]
    Returns: report_dates, date_urls, dates_checked
    """
    today = state.get("today") or date.today()
    log.info(f"[detect_dates] today={today}")

    # Build candidate dates: end-of-month window + mid-month window
    candidates: list[date] = []
    seen: set[date] = set()
    yr, mo = today.year, today.month

    for _ in range(14):                                  # ~7 months back
        last = calendar.monthrange(yr, mo)[1]
        for d in range(last, max(last - 3, 0), -1):     # end-of-month ±2
            dt = date(yr, mo, d)
            if dt <= today and dt not in seen:
                candidates.append(dt); seen.add(dt)
        for d in range(15, 12, -1):                     # mid-month ±2
            if d <= last:
                dt = date(yr, mo, d)
                if dt <= today and dt not in seen:
                    candidates.append(dt); seen.add(dt)
        mo -= 1
        if mo == 0: mo, yr = 12, yr - 1

    found:         list[date] = []
    date_urls:     dict[str, str] = {}
    dates_checked: list[str] = []
    session = requests.Session()
    session.headers.update(_HTTP_HEADERS)

    for d in candidates:
        if len(found) == 2:
            break
        u = _url(d)
        try:
            status = session.head(u, timeout=10).status_code
        except Exception as e:
            status = 0
            log.warning(f"  HEAD error {d}: {e}")

        dates_checked.append(f"{d} → HTTP {status}")
        log.info(f"  {d}  {u.split('/')[-1]}  HTTP {status}")

        if status == 200:
            found.append(d)
            date_urls[_ds(d)] = u
            log.info(f"  ✅ {d}")
        time.sleep(0.25)

    report_dates = sorted(found, reverse=True)
    log.info(f"[detect_dates] found={report_dates}")
    return {
        "report_dates":  report_dates,
        "date_urls":     date_urls,
        "dates_checked": dates_checked,
    }


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  NODE 2 — fetch_reports                                                      ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
def fetch_reports(state: FPIState) -> dict:
    """
    Download HTML for each date in report_dates.

    Reads  : state["report_dates"], state["date_urls"]
    Returns: raw_html, fetch_errors
    """
    report_dates = state.get("report_dates", [])
    date_urls    = state.get("date_urls", {})
    log.info(f"[fetch_reports] dates={report_dates}")

    raw_html:     dict[str, str] = {}
    fetch_errors: list[str] = []
    session = requests.Session()
    session.headers.update(_HTTP_HEADERS)

    for rd in report_dates:
        ds  = _ds(rd)
        url = date_urls.get(ds, _url(rd))
        try:
            r = session.get(url, timeout=30)
            if r.status_code == 200:
                raw_html[ds] = r.text
                log.info(f"  ✓ {ds}  ({len(r.text):,} bytes)")
            else:
                fetch_errors.append(f"{ds} HTTP {r.status_code}")
        except Exception as e:
            fetch_errors.append(f"{ds} error: {e}")
            log.error(f"  ✗ {ds}: {e}")
        time.sleep(1.0)

    return {"raw_html": raw_html, "fetch_errors": fetch_errors}


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  NODE 3 — parse_tables                                                       ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
def parse_tables(state: FPIState) -> dict:
    """
    Parse the main 98-col HTML table, skip 4 header rows.
    Stores raw rows as list[list[str]] per date.

    Reads  : state["raw_html"]
    Returns: raw_dataframes, parse_errors
    """
    raw_html = state.get("raw_html", {})
    log.info(f"[parse_tables] htmls={list(raw_html.keys())}")

    raw_dataframes: dict[str, list] = {}
    parse_errors:   list[str] = []

    for ds, html in raw_html.items():
        soup   = BeautifulSoup(html, "html.parser")
        tables = soup.find_all("table")
        if not tables:
            parse_errors.append(f"{ds}: no table found"); continue

        rows      = tables[0].find_all("tr")
        data_rows = []
        for row in rows[4:]:                             # skip 4 header rows
            cells = [c.get_text(" ", strip=True) for c in row.find_all(["td","th"])]
            if any(cells):
                data_rows.append((cells + [""] * 98)[:98])

        if not data_rows:
            parse_errors.append(f"{ds}: no data rows"); continue

        raw_dataframes[ds] = data_rows
        log.info(f"  ✓ {ds}: {len(data_rows)} rows")

    return {"raw_dataframes": raw_dataframes, "parse_errors": parse_errors}


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  NODE 4 — extract_data                                                       ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
def extract_data(state: FPIState) -> dict:
    """
    Map exact column indices → semantic field names for every sector row.

    Reads  : state["raw_dataframes"]
    Returns: sector_data, sector_totals, extract_errors
    """
    raw_dataframes = state.get("raw_dataframes", {})
    log.info(f"[extract_data] dfs={list(raw_dataframes.keys())}")

    sector_data:    dict[str, list[dict]] = {}
    sector_totals:  dict[str, dict] = {}
    extract_errors: list[str] = []

    for ds, rows in raw_dataframes.items():
        records: list[dict] = []
        totals = {k: 0.0 for k in ["F1_Total","F2_Total","Month_Total",
                                     "F1_Equity","F2_Equity","F1_Debt","F2_Debt"]}

        for row in rows:
            sector = str(row[_COL_SECTOR]).strip()
            if (not sector
                    or sector.lower() in ("","nan","total","grand total","sectors")
                    or sector.isdigit()):
                continue

            # ── Fortnight 1 ──────────────────────────────────────────────────
            f1_eq   = _num(row[_cidx("F1","Equity")])
            f1_debt = (_num(row[_cidx("F1","DebtGL")]) +
                       _num(row[_cidx("F1","DebtVRR")]) +
                       _num(row[_cidx("F1","DebtFAR")]))
            f1_hyb  = _num(row[_cidx("F1","Hybrid")])
            f1_mf   = _mf(row,"F1")
            f1_aif  = _num(row[_cidx("F1","AIF")])
            f1_tot  = _num(row[_cidx("F1","Total")])        # col 37
            f1_usd  = _num(row[_BLOCKS["F1"] + _USD_OFFSET + _SUB["Total"]])  # col 49

            # ── Fortnight 2 ──────────────────────────────────────────────────
            f2_eq   = _num(row[_cidx("F2","Equity")])
            f2_debt = (_num(row[_cidx("F2","DebtGL")]) +
                       _num(row[_cidx("F2","DebtVRR")]) +
                       _num(row[_cidx("F2","DebtFAR")]))
            f2_hyb  = _num(row[_cidx("F2","Hybrid")])
            f2_mf   = _mf(row,"F2")
            f2_aif  = _num(row[_cidx("F2","AIF")])
            f2_tot  = _num(row[_cidx("F2","Total")])        # col 61
            f2_usd  = _num(row[_BLOCKS["F2"] + _USD_OFFSET + _SUB["Total"]])  # col 73

            records.append({
                "sector":          sector,
                "date_str":        ds,
                # AUC
                "auc_f15_inr":     _num(row[_cidx("AUC_F15","Total")]),
                "auc_f28_inr":     _num(row[_cidx("AUC_F28","Total")]),
                # F1
                "f1_equity":       f1_eq,
                "f1_debt":         f1_debt,
                "f1_hybrid":       f1_hyb,
                "f1_mf":           f1_mf,
                "f1_aif":          f1_aif,
                "f1_total_inr":    f1_tot,
                "f1_total_usd":    f1_usd,
                # F2
                "f2_equity":       f2_eq,
                "f2_debt":         f2_debt,
                "f2_hybrid":       f2_hyb,
                "f2_mf":           f2_mf,
                "f2_aif":          f2_aif,
                "f2_total_inr":    f2_tot,
                "f2_total_usd":    f2_usd,
                # Month
                "month_equity":    f1_eq  + f2_eq,
                "month_debt":      f1_debt + f2_debt,
                "month_hybrid":    f1_hyb + f2_hyb,
                "month_mf":        f1_mf  + f2_mf,
                "month_aif":       f1_aif + f2_aif,
                "month_total_inr": f1_tot + f2_tot,
                "month_total_usd": f1_usd + f2_usd,
            })

            totals["F1_Total"]   += f1_tot
            totals["F2_Total"]   += f2_tot
            totals["Month_Total"]+= f1_tot + f2_tot
            totals["F1_Equity"]  += f1_eq
            totals["F2_Equity"]  += f2_eq
            totals["F1_Debt"]    += f1_debt
            totals["F2_Debt"]    += f2_debt

        sector_data[ds]  = records
        sector_totals[ds] = totals
        log.info(f"  ✓ {ds}: {len(records)} sectors | Net ₹{totals['Month_Total']:,.0f} Cr")

    return {
        "sector_data":    sector_data,
        "sector_totals":  sector_totals,
        "extract_errors": extract_errors,
    }


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  NODE 5 — generate_insights                                                  ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
def generate_insights(state: FPIState) -> dict:
    """
    Cross-period signals, momentum, buy/sell classification, top movers.

    Reads  : state["sector_data"], state["sector_totals"], state["report_dates"]
    Returns: insights, summary, buying_sectors, selling_sectors, top_movers
    """
    sector_data   = state.get("sector_data", {})
    sector_totals = state.get("sector_totals", {})
    log.info(f"[generate_insights] dates={list(sector_data.keys())}")

    if not sector_data:
        return {"insights":[],"summary":{},"buying_sectors":[],
                "selling_sectors":[],"top_movers":[]}

    dates_sorted = sorted(sector_data.keys(), reverse=True)
    ds_new = dates_sorted[0]
    ds_old = dates_sorted[1] if len(dates_sorted) > 1 else None

    new_by_sector = {r["sector"]: r for r in sector_data[ds_new]}
    old_by_sector = {r["sector"]: r for r in sector_data[ds_old]} if ds_old else {}

    insights:        list[dict] = []
    buying_sectors:  list[dict] = []
    selling_sectors: list[dict] = []

    for sector, nr in new_by_sector.items():
        or_ = old_by_sector.get(sector, {})
        f1, f2   = nr["f1_total_inr"], nr["f2_total_inr"]
        month    = nr["month_total_inr"]
        prev     = or_.get("month_total_inr", 0.0)
        change   = month - prev

        rec = {
            "sector":          sector,
            "date_new":        ds_new,
            "date_old":        ds_old or "",
            # net investment
            "f1_net_inr":      f1,
            "f2_net_inr":      f2,
            "month_net_inr":   month,
            # by asset class
            "month_equity":    nr["month_equity"],
            "month_debt":      nr["month_debt"],
            "month_hybrid":    nr["month_hybrid"],
            "month_mf":        nr["month_mf"],
            "month_aif":       nr["month_aif"],
            # vs prev report
            "prev_month_net":  prev,
            "change_vs_prev":  change,
            # signals
            "signal_f1":       _signal(f1),
            "signal_f2":       _signal(f2),
            "signal_month":    _signal(month),
            "signal_equity":   _signal(nr["month_equity"]),
            "signal_debt":     _signal(nr["month_debt"]),
            # qualitative
            "momentum":       ("ACCELERATING ▲" if f2 > f1 else
                               "DECELERATING ▼" if f2 < f1 else "STABLE →"),
            "asset_dominance":("Equity-led"
                               if abs(nr["month_equity"]) >= abs(nr["month_debt"])
                               else "Debt-led"),
            # AUC
            "auc_f15_inr":    nr.get("auc_f15_inr", 0),
            "auc_f28_inr":    nr.get("auc_f28_inr", 0),
        }
        insights.append(rec)
        if month > 0:  buying_sectors.append(rec)
        elif month < 0: selling_sectors.append(rec)

    insights.sort(       key=lambda x: x["month_net_inr"], reverse=True)
    buying_sectors.sort( key=lambda x: x["month_net_inr"], reverse=True)
    selling_sectors.sort(key=lambda x: x["month_net_inr"])
    top_movers = sorted(insights, key=lambda x: abs(x["change_vs_prev"]), reverse=True)[:5]

    # ── overall market summary ────────────────────────────────────────────────
    nt = sector_totals.get(ds_new, {})
    ot = sector_totals.get(ds_old, {}) if ds_old else {}
    month_total = nt.get("Month_Total", 0)
    prev_total  = ot.get("Month_Total", 0)

    summary = {
        "date_new":          ds_new,
        "date_old":          ds_old or "",
        "total_f1_inr":      nt.get("F1_Total", 0),
        "total_f2_inr":      nt.get("F2_Total", 0),
        "total_month_inr":   month_total,
        "total_equity_inr":  nt.get("F1_Equity",0) + nt.get("F2_Equity",0),
        "total_debt_inr":    nt.get("F1_Debt",0)   + nt.get("F2_Debt",0),
        "prev_month_inr":    prev_total,
        "change_vs_prev":    month_total - prev_total,
        "overall_signal":    _signal(month_total),
        "n_buying":          len(buying_sectors),
        "n_selling":         len(selling_sectors),
        "top_buy_sector":    buying_sectors[0]["sector"]  if buying_sectors  else "N/A",
        "top_sell_sector":   selling_sectors[0]["sector"] if selling_sectors else "N/A",
    }

    _print_insights(insights, buying_sectors, selling_sectors, top_movers, summary)

    return {
        "insights":        insights,
        "summary":         summary,
        "buying_sectors":  buying_sectors,
        "selling_sectors": selling_sectors,
        "top_movers":      top_movers,
    }


def _print_insights(insights, buying, selling, movers, summary):
    W = 108
    print(f"\n{'═'*W}")
    print(f"  📊  FPI INSIGHTS  |  {summary['date_old']} → {summary['date_new']}")
    print(f"{'═'*W}")
    print(f"  F1: ₹{summary['total_f1_inr']:>+12,.0f} Cr   "
          f"F2: ₹{summary['total_f2_inr']:>+12,.0f} Cr   "
          f"Month: ₹{summary['total_month_inr']:>+12,.0f} Cr   "
          f"Signal: {summary['overall_signal']}")
    print(f"  Equity: ₹{summary['total_equity_inr']:>+12,.0f} Cr   "
          f"Debt: ₹{summary['total_debt_inr']:>+12,.0f} Cr   "
          f"Change vs Prev: ₹{summary['change_vs_prev']:>+10,.0f} Cr")
    print(f"  Buying: {summary['n_buying']} sectors  |  Selling: {summary['n_selling']} sectors")

    print(f"\n{'─'*W}")
    print(f"  {'#':>2}  {'Sector':<38}  {'F1':>10}  {'F2':>10}  "
          f"{'Month':>11}  {'Signal':<12}  {'Momentum':<17}  Asset")
    print(f"{'─'*W}")
    for i, r in enumerate(insights, 1):
        print(f"  {i:>2}  {r['sector'][:38]:<38}  {r['f1_net_inr']:>+10,.0f}  "
              f"{r['f2_net_inr']:>+10,.0f}  {r['month_net_inr']:>+11,.0f}  "
              f"{r['signal_month']:<12}  {r['momentum']:<17}  {r['asset_dominance']}")

    print(f"\n{'═'*W}")
    print(f"  🟢  BUYING ({len(buying)} sectors)")
    print(f"{'─'*W}")
    for r in buying:
        print(f"  {r['sector']:<42}  ₹{r['month_net_inr']:>+12,.0f}  "
              f"Eq:{r['month_equity']:>+10,.0f}  Debt:{r['month_debt']:>+10,.0f}  {r['momentum']}")

    print(f"\n{'═'*W}")
    print(f"  🔴  SELLING ({len(selling)} sectors)")
    print(f"{'─'*W}")
    for r in selling:
        print(f"  {r['sector']:<42}  ₹{r['month_net_inr']:>+12,.0f}  "
              f"Eq:{r['month_equity']:>+10,.0f}  Debt:{r['month_debt']:>+10,.0f}  {r['momentum']}")

    print(f"\n{'═'*W}")
    print(f"  🔄  TOP 5 MOVERS vs PREVIOUS REPORT")
    print(f"{'─'*W}")
    for r in movers:
        print(f"  {r['sector']:<42}  Change: ₹{r['change_vs_prev']:>+12,.0f}  "
              f"{'▲' if r['change_vs_prev']>=0 else '▼'}  ({r['signal_month']})")
    print(f"{'═'*W}\n")


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  NODE 6 — save_reports                                                       ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
def save_reports(state: FPIState) -> dict:
    """
    Write colour-coded Excel + individual CSVs from state.

    Reads  : state["insights"], state["buying_sectors"], state["selling_sectors"],
             state["top_movers"], state["summary"], state["sector_data"],
             state["output_dir"]
    Returns: saved_files, excel_path, save_errors
    """
    from openpyxl.styles import PatternFill, Font, Alignment

    output_dir = state.get("output_dir", "fpi_data")
    insights   = state.get("insights", [])
    buying     = state.get("buying_sectors", [])
    selling    = state.get("selling_sectors", [])
    movers     = state.get("top_movers", [])
    summary    = state.get("summary", {})
    sector_data= state.get("sector_data", {})
    ds_new     = summary.get("date_new", "unknown")
    log.info(f"[save_reports] dir={output_dir}  date={ds_new}")

    os.makedirs(output_dir, exist_ok=True)
    saved_files: list[str] = []
    save_errors: list[str] = []

    def _df(records): return pd.DataFrame(records) if records else pd.DataFrame()

    # ── CSVs ──────────────────────────────────────────────────────────────────
    for fname, data in {
        f"FPI_Insights_{ds_new}.csv":  insights,
        f"FPI_Buying_{ds_new}.csv":    buying,
        f"FPI_Selling_{ds_new}.csv":   selling,
        f"FPI_TopMovers_{ds_new}.csv": movers,
        f"FPI_Summary_{ds_new}.csv":   [summary] if summary else [],
    }.items():
        if data:
            p = os.path.join(output_dir, fname)
            _df(data).to_csv(p, index=False)
            saved_files.append(os.path.abspath(p))

    for ds, records in sector_data.items():
        p = os.path.join(output_dir, f"FPI_Raw_{ds}.csv")
        _df(records).to_csv(p, index=False)
        saved_files.append(os.path.abspath(p))

    # ── Excel ──────────────────────────────────────────────────────────────────
    xl = os.path.join(output_dir, f"FPI_Report_{ds_new}.xlsx")
    try:
        def _style(ws, df, sig_col=None):
            for cell in ws[1]:
                cell.fill      = PatternFill("solid", fgColor="2C0066")
                cell.font      = Font(color="FFFFFF", bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            sig_ci = (list(df.columns).index(sig_col) + 1) if sig_col and sig_col in df.columns else None

            for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
                sig = str(ws.cell(ri, sig_ci).value or "") if sig_ci else ""
                txt_c = bg_c = None
                for key, (tc, bc) in _SIGNAL_COLORS.items():
                    if key in sig: txt_c, bg_c = tc, bc; break

                for cell in row:
                    if bg_c: cell.fill = PatternFill("solid", fgColor=bg_c)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0.00"
                        color = txt_c or ("CC0000" if cell.value < 0 else "006400")
                        cell.font = Font(color=color, bold=bool(txt_c))

            for col_cells in ws.columns:
                w = max((len(str(c.value or "")) for c in col_cells), default=8)
                ws.column_dimensions[col_cells[0].column_letter].width = min(w + 3, 44)
            ws.freeze_panes = "B2"

        with pd.ExcelWriter(xl, engine="openpyxl") as writer:
            sheets = [
                ("📊 Insights",    _df(insights), "signal_month"),
                ("🟢 Buying",      _df(buying),   "signal_month"),
                ("🔴 Selling",     _df(selling),  "signal_month"),
                ("🔄 Top Movers",  _df(movers),   "signal_month"),
                ("📋 Summary",     _df([summary] if summary else []), None),
            ]
            for ds, records in sector_data.items():
                sheets.append((f"Raw_{ds}", _df(records), None))

            for name, df, sig_col in sheets:
                if df.empty: continue
                df.to_excel(writer, sheet_name=name[:31], index=False)
                _style(writer.sheets[name[:31]], df, sig_col)

        saved_files.append(os.path.abspath(xl))
        log.info(f"  💾 {xl}")

    except Exception as e:
        save_errors.append(f"Excel: {e}")
        log.error(f"  ✗ Excel: {e}")

    print(f"\n  ✅ Saved {len(saved_files)} files → {os.path.abspath(output_dir)}")
    return {
        "saved_files": saved_files,
        "excel_path":  os.path.abspath(xl),
        "save_errors": save_errors,
    }


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  SEQUENTIAL RUNNER — no LangGraph needed                                     ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
def run(today: date = None, output_dir: str = "fpi_data") -> FPIState:
    """Run all nodes in order, passing state through each step."""

    # ── initial state ─────────────────────────────────────────────────────────
    state: FPIState = {
        "today":      today or date.today(),
        "output_dir": output_dir,
    }

    print(f"\n{'═'*60}")
    print(f"  FPI Pipeline  |  {state['today'].strftime('%d-%b-%Y')}")
    print(f"{'═'*60}")

    # ── run each node, merge its output back into state ───────────────────────
    for node_fn in [detect_dates, fetch_reports, parse_tables,
                    extract_data, generate_insights, save_reports]:
        print(f"\n▶  {node_fn.__name__}")
        state.update(node_fn(state))

        # stop early if no reports were found
        if node_fn.__name__ == "detect_dates" and not state.get("report_dates"):
            print("  ✗ No reports found — stopping.")
            break

    return state


if __name__ == "__main__":
    final_state = run()
